#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==========================================================================
  公司名称智能匹配系统 — Web Edition
  Company Name Intelligent Matching System — Flask Web Application
==========================================================================
  基于原始 Python 匹配引擎，封装为 Flask Web 应用
  端口: 8080
  运行: python app.py
==========================================================================
"""

import os
import sys
import re
import math
import json
import uuid
import time
import threading
from collections import Counter
from datetime import datetime

# ======================== 第三方库导入 ========================
try:
    from flask import (Flask, render_template_string, request,
                       jsonify, send_file, session)
except ImportError:
    print("[!] 正在安装 Flask...")
    os.system(f"{sys.executable} -m pip install flask -q")
    from flask import (Flask, render_template_string, request,
                       jsonify, send_file, session)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("[!] 正在安装 openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import pandas as pd
except ImportError:
    print("[!] 正在安装 pandas...")
    os.system(f"{sys.executable} -m pip install pandas -q")
    import pandas as pd

try:
    from opencc import OpenCC
    CC_T2S = OpenCC('t2s')
    HAS_OPENCC = True
except ImportError:
    try:
        os.system(f"{sys.executable} -m pip install opencc-python-reimplemented -q")
        from opencc import OpenCC
        CC_T2S = OpenCC('t2s')
        HAS_OPENCC = True
    except Exception:
        HAS_OPENCC = False

# ======================== Flask App ========================
app = Flask(__name__)
app.secret_key = 'company-matcher-secret-key-2024'

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
RESULT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'results')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

# 全局任务状态存储
tasks = {}

# ======================== 繁简体映射（备用） ========================
TRAD_TO_SIMP = {
    '國': '国', '際': '际', '業': '业', '門': '门', '開': '开', '發': '发',
    '會': '会', '經': '经', '濟': '济', '貿': '贸', '電': '电', '機': '机',
    '設': '设', '計': '计', '環': '环', '達': '达', '運': '运', '輸': '输',
    '銀': '银', '號': '号', '萬': '万', '億': '亿', '華': '华', '東': '东',
    '車': '车', '馬': '马', '龍': '龙', '鳳': '凤', '聯': '联', '豐': '丰',
    '廣': '广', '產': '产', '實': '实', '寶': '宝', '與': '与', '對': '对',
    '長': '长', '創': '创', '風': '风', '園': '园', '築': '筑', '買': '买',
    '賣': '卖', '進': '进', '農': '农', '漁': '渔', '礦': '矿', '鋼': '钢',
    '鐵': '铁', '飛': '飞', '樂': '乐', '書': '书', '學': '学', '類': '类',
    '師': '师', '個': '个', '從': '从', '復': '复', '後': '后', '態': '态',
    '總': '总', '團': '团', '層': '层', '陽': '阳', '島': '岛', '區': '区',
    '織': '织', '紡': '纺', '線': '线', '練': '练', '組': '组', '終': '终',
    '約': '约', '級': '级', '結': '结', '給': '给', '統': '统', '網': '网',
    '護': '护', '報': '报', '場': '场', '塊': '块', '壓': '压', '夢': '梦',
    '備': '备', '處': '处', '複': '复', '裝': '装', '規': '规', '視': '视',
    '覺': '觉', '記': '记', '許': '许', '論': '论', '證': '证', '識': '识',
    '議': '议', '財': '财', '質': '质', '貨': '货', '資': '资', '賢': '贤',
    '軟': '软', '輝': '辉', '辦': '办', '邊': '边', '遠': '远', '選': '选',
    '鑫': '鑫', '關': '关', '陳': '陈', '項': '项', '順': '顺', '領': '领',
    '題': '题', '點': '点', '傳': '传', '僅': '仅', '優': '优', '衛': '卫',
    '寫': '写', '導': '导', '盡': '尽', '屬': '属', '嶺': '岭', '幣': '币',
    '廳': '厅', '張': '张', '彈': '弹', '強': '强', '歸': '归', '當': '当',
    '徑': '径', '應': '应', '戲': '戏', '擴': '扩', '據': '据',
    '損': '损', '搖': '摇', '構': '构', '標': '标', '歐': '欧', '歷': '历',
    '殘': '残', '滅': '灭', '瀾': '澜', '灣': '湾', '為': '为', '無': '无',
    '營': '营', '獨': '独', '獎': '奖', '現': '现', '瑣': '琐', '畫': '画',
    '異': '异', '療': '疗', '盤': '盘', '監': '监', '碼': '码', '積': '积',
    '稅': '税', '穩': '稳', '競': '竞', '節': '节', '範': '范',
    '係': '系', '紀': '纪', '蘭': '兰', '蘇': '苏', '術': '术', '衝': '冲',
    '補': '补', '製': '制', '覆': '覆', '親': '亲', '觸': '触',
    '訊': '讯', '調': '调', '談': '谈', '請': '请', '變': '变',
    '讓': '让', '豬': '猪', '負': '负', '販': '贩', '費': '费', '離': '离',
    '雲': '云', '響': '响', '頭': '头', '顯': '显', '養': '养', '體': '体',
    '魚': '鱼', '麗': '丽', '齊': '齐', '齡': '龄',
}

# ======================== 文本标准化引擎 ========================

def traditional_to_simplified(text: str) -> str:
    if HAS_OPENCC:
        return CC_T2S.convert(text)
    return ''.join(TRAD_TO_SIMP.get(c, c) for c in text)

SUFFIX_NORMALIZE = [
    (r'\blimited\b', 'ltd'), (r'\bltd\.?\b', 'ltd'),
    (r'\bcorporation\b', 'corp'), (r'\bcorp\.?\b', 'corp'),
    (r'\bcompany\b', 'co'), (r'\bco\.?\b', 'co'),
    (r'\bincorporated\b', 'inc'), (r'\binc\.?\b', 'inc'),
    (r'\bholdings?\b', 'holding'), (r'\bgroup\b', 'group'),
    (r'\binternational\b', 'intl'), (r'\bintl\.?\b', 'intl'),
    (r'\benterprise[s]?\b', 'enterprise'),
    (r'\binvestment[s]?\b', 'investment'),
    (r'\btechnolog(?:y|ies)\b', 'tech'), (r'\btech\.?\b', 'tech'),
    (r'\bmanagement\b', 'mgmt'), (r'\bmgmt\.?\b', 'mgmt'),
    (r'\bdevelopment\b', 'dev'), (r'\bdev\.?\b', 'dev'),
    (r'\bprivate\b', 'pvt'), (r'\bpvt\.?\b', 'pvt'),
    (r'\bsdn\.?\s*bhd\.?\b', 'sdn bhd'), (r'\bpte\.?\b', 'pte'),
    (r'\bb\.?\s*v\.?\b', 'bv'), (r'\bn\.?\s*v\.?\b', 'nv'),
    (r'\bg\.?\s*m\.?\s*b\.?\s*h\.?\b', 'gmbh'),
    (r'\bs\.?\s*a\.?\s*r\.?\s*l\.?\b', 'sarl'),
    (r'\bs\.?\s*\.?a\.?\b', 'sa'),
    (r'有限责任公司', '有限公司'), (r'股份有限公司', '有限公司'),
    (r'（香港）', '(香港)'), (r'（中国）', '(中国)'),
    (r'（上海）', '(上海)'), (r'（北京）', '(北京)'),
    (r'（深圳）', '(深圳)'), (r'（广州）', '(广州)'),
]

STOP_WORDS = {
    'the', 'of', 'and', '&', 'a', 'an', 'in', 'at', 'on', 'for', 'to', 'by',
    '-', '—', '–', '·', '•', ',', '.', '/', '\\',
}


def full_to_half(text: str) -> str:
    result = []
    for char in text:
        code = ord(char)
        if 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        elif code == 0x3000:
            result.append(' ')
        else:
            result.append(char)
    return ''.join(result)


def normalize_company_name(name: str) -> str:
    if not name or not isinstance(name, str):
        return ''
    text = name.strip()
    text = traditional_to_simplified(text)
    text = full_to_half(text)
    text = text.lower()
    for pattern, replacement in SUFFIX_NORMALIZE:
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    text = text.replace('（', '(').replace('）', ')')
    text = text.replace('【', '(').replace('】', ')')
    text = text.replace('〔', '(').replace('〕', ')')
    text = text.replace('［', '(').replace('］', ')')
    text = text.replace('｛', '(').replace('｝', ')')
    text = text.replace('{', '(').replace('}', ')')
    text = text.replace('[', '(').replace(']', ')')
    text = text.replace('﹙', '(').replace('﹚', ')')
    text = text.replace('《', '(').replace('》', ')')
    text = text.replace('，', ',').replace('。', '.').replace('；', ';')
    text = text.replace('：', ':').replace('、', ',').replace('～', '~')
    text = text.replace('\u3000', ' ').replace('\xa0', ' ')
    text = text.replace('\t', ' ').replace('\r', ' ').replace('\n', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def extract_chars(text: str) -> list:
    return [c for c in text if c.isalnum() or '\u4e00' <= c <= '\u9fff']


def extract_tokens(text: str) -> list:
    tokens = []
    parts = text.split()
    for part in parts:
        if part.lower() in STOP_WORDS:
            continue
        i = 0
        current_eng = []
        while i < len(part):
            if '\u4e00' <= part[i] <= '\u9fff':
                if current_eng:
                    eng_word = ''.join(current_eng)
                    tokens.append(eng_word)
                    if len(eng_word) >= 3:
                        tokens.extend(list(eng_word))
                    current_eng = []
                tokens.append(part[i])
            elif part[i].isalnum():
                current_eng.append(part[i])
            else:
                if current_eng:
                    eng_word = ''.join(current_eng)
                    tokens.append(eng_word)
                    if len(eng_word) >= 3:
                        tokens.extend(list(eng_word))
                    current_eng = []
            i += 1
        if current_eng:
            eng_word = ''.join(current_eng)
            tokens.append(eng_word)
            if len(eng_word) >= 3:
                tokens.extend(list(eng_word))
    return tokens


# ======================== 相似度计算引擎 ========================

def char_overlap_ratio(chars_a, chars_b):
    if not chars_a or not chars_b:
        return 0.0
    counter_a = Counter(chars_a)
    counter_b = Counter(chars_b)
    hit_b_in_a = sum(min(counter_b[c], counter_a.get(c, 0)) for c in counter_b)
    hit_a_in_b = sum(min(counter_a[c], counter_b.get(c, 0)) for c in counter_a)
    ratio_b = hit_b_in_a / len(chars_b) if chars_b else 0
    ratio_a = hit_a_in_b / len(chars_a) if chars_a else 0
    return 0.6 * ratio_b + 0.4 * ratio_a


def longest_common_subsequence_ratio(s1, s2):
    if not s1 or not s2:
        return 0.0
    m, n = len(s1), len(s2)
    prev = [0] * (n + 1)
    curr = [0] * (n + 1)
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if s1[i-1] == s2[j-1]:
                curr[j] = prev[j-1] + 1
            else:
                curr[j] = max(prev[j], curr[j-1])
        prev, curr = curr, [0] * (n + 1)
    lcs_len = prev[n]
    return (2.0 * lcs_len) / (m + n)


def token_overlap_ratio(tokens_a, tokens_b):
    if not tokens_a or not tokens_b:
        return 0.0
    set_a = Counter(tokens_a)
    set_b = Counter(tokens_b)
    hit = sum(min(set_a.get(t, 0), set_b[t]) for t in set_b)
    total = max(len(tokens_a), len(tokens_b))
    return hit / total if total > 0 else 0.0


def compute_similarity(name_a_norm, name_b_norm, chars_a, chars_b, tokens_a, tokens_b):
    if name_a_norm == name_b_norm:
        return 1.0
    s1 = char_overlap_ratio(chars_a, chars_b)
    s2 = longest_common_subsequence_ratio(name_a_norm, name_b_norm)
    s3 = token_overlap_ratio(tokens_a, tokens_b)
    return 0.40 * s1 + 0.35 * s2 + 0.25 * s3


# ======================== 核心匹配流程 ========================

def read_company_list(filepath, scan_all_cells=False):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        df = pd.read_excel(filepath, header=None, dtype=str)
    elif ext == '.csv':
        df = pd.read_csv(filepath, header=None, dtype=str)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")

    if scan_all_cells:
        names_set = []
        seen = set()
        for col in df.columns:
            for val in df[col].dropna().astype(str):
                val = val.strip()
                if val and val not in seen:
                    if not re.fullmatch(r'[-+]?\d*\.?\d+', val):
                        names_set.append(val)
                        seen.add(val)
        return names_set
    else:
        names = df.iloc[:, 0].dropna().astype(str).tolist()
        names = [n.strip() for n in names if n.strip()]
        return names


def detect_table_format(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        df = pd.read_excel(filepath, header=None, dtype=str)
    elif ext == '.csv':
        df = pd.read_csv(filepath, header=None, dtype=str)
    else:
        return True, 0, 0

    total_rows = len(df)
    total_cols = len(df.columns)

    if df.shape[1] <= 1:
        count = df.iloc[:, 0].dropna().astype(str).apply(lambda x: len(x.strip()) > 0).sum() if total_cols > 0 else 0
        return True, total_rows, count

    col_fill_rates = []
    for col in df.columns:
        non_empty = df[col].dropna().astype(str).apply(lambda x: len(x.strip()) > 0).sum()
        col_fill_rates.append(non_empty / total_rows if total_rows > 0 else 0)

    first_col_rate = col_fill_rates[0]
    other_cols_rate = sum(col_fill_rates[1:]) / max(len(col_fill_rates) - 1, 1)

    # Count non-empty cells for display
    first_col_count = df.iloc[:, 0].dropna().astype(str).apply(lambda x: len(x.strip()) > 0).sum()

    if first_col_rate > 0.8 and other_cols_rate < 0.2:
        return True, total_rows, first_col_count
    if sum(1 for r in col_fill_rates if r > 0.3) > 1:
        # Count all unique non-empty text cells
        all_count = 0
        seen = set()
        for col in df.columns:
            for val in df[col].dropna().astype(str):
                val = val.strip()
                if val and val not in seen and not re.fullmatch(r'[-+]?\d*\.?\d+', val):
                    all_count += 1
                    seen.add(val)
        return False, total_rows, all_count
    return True, total_rows, first_col_count


def build_index(names):
    index = []
    for name in names:
        norm = normalize_company_name(name)
        chars = extract_chars(norm)
        tokens = extract_tokens(norm)
        index.append((name, norm, chars, tokens))
    return index


def match_all_with_progress(index_a, index_b, top_pct, task_id):
    """带进度跟踪的匹配"""
    total_a = len(index_a)
    top_rank_limit = max(1, math.ceil(total_a * top_pct))
    best_matches = []
    detail_matches = []
    total_b = len(index_b)

    for idx_b, (b_name, b_norm, b_chars, b_tokens) in enumerate(index_b):
        # 更新进度
        progress = (idx_b + 1) / total_b * 100
        tasks[task_id]['progress'] = progress
        tasks[task_id]['current'] = idx_b + 1
        tasks[task_id]['current_name'] = b_name[:30] + '...' if len(b_name) > 30 else b_name

        scores = []
        for (a_name, a_norm, a_chars, a_tokens) in index_a:
            sim = compute_similarity(a_norm, b_norm, a_chars, b_chars, a_tokens, b_tokens)
            scores.append((a_name, sim))

        scores.sort(key=lambda x: -x[1])
        best_a_name, best_score = scores[0] if scores else ('', 0.0)
        best_matches.append((b_name, best_a_name, best_score))

        top_list = []
        current_rank = 0
        prev_score = None
        count = 0
        for a_name, score in scores:
            count += 1
            if score != prev_score:
                current_rank = count
                prev_score = score
            if current_rank <= top_rank_limit:
                top_list.append((a_name, score))
            else:
                break
        detail_matches.append((b_name, top_list))

    return best_matches, detail_matches


# ======================== Excel 输出 ========================

def write_best_match_sheet(wb, best_matches, detail_matches, threshold=0.35):
    ws = wb.active
    ws.title = "最佳匹配结果"
    detail_lookup = {b_name: top_list for b_name, top_list in detail_matches}

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["B表原始公司名称", "A表最佳匹配公司", "匹配度", "风险提示", "前X%候选公司明细"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    for row_idx, (b_name, a_name, score) in enumerate(best_matches, 2):
        ws.cell(row=row_idx, column=1, value=b_name).border = thin_border
        ws.cell(row=row_idx, column=2, value=a_name).border = thin_border

        score_cell = ws.cell(row=row_idx, column=3, value=round(score * 100, 2))
        score_cell.number_format = '0.00"%"'
        score_cell.border = thin_border
        score_cell.alignment = Alignment(horizontal="center")

        note_cell = ws.cell(row=row_idx, column=4)
        note_cell.border = thin_border

        if score >= 0.85:
            score_cell.fill = green_fill
            note_cell.value = "✓ 高度匹配"
            note_cell.font = Font(color="006100")
        elif score >= 0.55:
            score_cell.fill = yellow_fill
            note_cell.value = "⚠ 中等匹配，建议人工复核"
            note_cell.font = Font(color="9C6500")
        elif score >= threshold:
            score_cell.fill = red_fill
            note_cell.value = "⚠ 低匹配度，请重点核查"
            note_cell.font = Font(color="9C0006")
        else:
            score_cell.fill = red_fill
            note_cell.value = "✗ 未找到相似公司"
            note_cell.font = Font(color="9C0006", bold=True)

        candidate_cell = ws.cell(row=row_idx, column=5)
        candidate_cell.border = thin_border
        candidate_cell.alignment = Alignment(vertical="top", wrap_text=True)

        if score < 1.0:
            top_list = detail_lookup.get(b_name, [])
            if top_list:
                lines = []
                rank = 0
                prev_s = None
                for a_candidate, s in top_list:
                    if s != prev_s:
                        rank += 1
                        prev_s = s
                    lines.append(f"#{rank} [{s*100:.1f}%] {a_candidate}")
                candidate_cell.value = "\n".join(lines)
                candidate_cell.font = Font(size=9, name="Arial")

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 65
    ws.freeze_panes = 'A2'


def write_detail_sheet(wb, detail_matches, top_pct):
    ws = wb.create_sheet(title=f"前{int(top_pct*100)}%匹配明细")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    section_font = Font(bold=True, size=11, color="2F5496", name="Arial")
    section_fill = PatternFill("solid", fgColor="D6E4F0")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    main_headers = ["B表公司", "排名", "A表匹配公司", "匹配度(%)"]
    for col, h in enumerate(main_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    current_row = 2
    for b_name, top_list in detail_matches:
        b_cell = ws.cell(row=current_row, column=1, value=b_name)
        b_cell.font = section_font
        b_cell.fill = section_fill
        b_cell.border = thin_border

        if not top_list:
            ws.cell(row=current_row, column=2, value="-").border = thin_border
            ws.cell(row=current_row, column=3, value="无匹配结果").border = thin_border
            ws.cell(row=current_row, column=4, value=0).border = thin_border
            current_row += 1
            continue

        rank = 0
        prev_score = None
        for i, (a_name, score) in enumerate(top_list):
            if score != prev_score:
                rank = i + 1
                prev_score = score
            row = current_row + i
            if i > 0:
                ws.cell(row=row, column=1, value="").border = thin_border
            ws.cell(row=row, column=2, value=rank).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=a_name).border = thin_border
            score_cell = ws.cell(row=row, column=4, value=round(score * 100, 2))
            score_cell.number_format = '0.00'
            score_cell.border = thin_border
            score_cell.alignment = Alignment(horizontal="center")
        current_row += len(top_list)

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 14
    ws.freeze_panes = 'A2'


# ======================== 后台任务处理 ========================

def run_matching_task(task_id, path_a, path_b, scan_a, scan_b, top_pct):
    """在后台线程中执行匹配任务"""
    try:
        tasks[task_id]['status'] = 'reading'
        tasks[task_id]['message'] = '正在读取A表数据...'

        list_a = read_company_list(path_a, scan_all_cells=scan_a)
        tasks[task_id]['count_a'] = len(list_a)
        tasks[task_id]['message'] = f'A表读取完成：{len(list_a)} 家公司'

        list_b = read_company_list(path_b, scan_all_cells=scan_b)
        tasks[task_id]['count_b'] = len(list_b)
        tasks[task_id]['message'] = f'B表读取完成：{len(list_b)} 家公司'

        tasks[task_id]['status'] = 'indexing'
        tasks[task_id]['message'] = '正在构建预处理索引...'
        index_a = build_index(list_a)
        index_b = build_index(list_b)

        tasks[task_id]['status'] = 'matching'
        tasks[task_id]['total'] = len(list_b)
        tasks[task_id]['total_comparisons'] = len(list_a) * len(list_b)
        tasks[task_id]['message'] = f'开始匹配：{len(list_b)} × {len(list_a)} = {len(list_b)*len(list_a):,} 次比较'

        best_matches, detail_matches = match_all_with_progress(index_a, index_b, top_pct, task_id)

        tasks[task_id]['status'] = 'saving'
        tasks[task_id]['message'] = '正在生成Excel报告...'

        output_filename = f"匹配结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(RESULT_FOLDER, output_filename)

        wb = openpyxl.Workbook()
        write_best_match_sheet(wb, best_matches, detail_matches)
        write_detail_sheet(wb, detail_matches, top_pct)
        wb.save(output_path)

        # 统计
        high = sum(1 for _, _, s in best_matches if s >= 0.85)
        mid = sum(1 for _, _, s in best_matches if 0.55 <= s < 0.85)
        low = sum(1 for _, _, s in best_matches if s < 0.55)

        # 准备预览数据
        preview = []
        for b_name, a_name, score in best_matches[:100]:
            level = 'high' if score >= 0.85 else ('mid' if score >= 0.55 else 'low')
            preview.append({
                'b_name': b_name,
                'a_name': a_name,
                'score': round(score * 100, 2),
                'level': level
            })

        tasks[task_id]['status'] = 'done'
        tasks[task_id]['message'] = '匹配完成！'
        tasks[task_id]['progress'] = 100
        tasks[task_id]['result'] = {
            'filename': output_filename,
            'high': high,
            'mid': mid,
            'low': low,
            'total': len(best_matches),
            'preview': preview
        }

    except Exception as e:
        tasks[task_id]['status'] = 'error'
        tasks[task_id]['message'] = f'错误: {str(e)}'


# ======================== HTML 模板 ========================

HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NameLink — 公司名称智能匹配系统</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700&family=JetBrains+Mono:wght@400;500&family=Noto+Sans+SC:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
*, *::before, *::after { margin:0; padding:0; box-sizing:border-box; }

:root {
  --bg: #FAF9F7;
  --bg-card: #FFFFFF;
  --bg-subtle: #F3F1EE;
  --bg-hover: #ECEAE6;
  --text: #1A1A1A;
  --text-secondary: #6B6560;
  --text-tertiary: #9C9792;
  --accent: #2563EB;
  --accent-light: #DBEAFE;
  --accent-hover: #1D4ED8;
  --border: #E5E2DD;
  --border-strong: #D1CEC8;
  --green: #059669;
  --green-bg: #ECFDF5;
  --yellow: #D97706;
  --yellow-bg: #FFFBEB;
  --red: #DC2626;
  --red-bg: #FEF2F2;
  --shadow-sm: 0 1px 2px rgba(0,0,0,0.04);
  --shadow-md: 0 4px 16px rgba(0,0,0,0.06);
  --shadow-lg: 0 12px 40px rgba(0,0,0,0.08);
  --radius: 12px;
  --radius-sm: 8px;
  --radius-lg: 16px;
}

html { scroll-behavior: smooth; }

body {
  font-family: 'DM Sans', 'Noto Sans SC', -apple-system, sans-serif;
  background: var(--bg);
  color: var(--text);
  line-height: 1.6;
  min-height: 100vh;
  -webkit-font-smoothing: antialiased;
}

/* ===== NOISE TEXTURE OVERLAY ===== */
body::before {
  content: '';
  position: fixed;
  top: 0; left: 0; right: 0; bottom: 0;
  background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='noise'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23noise)' opacity='0.03'/%3E%3C/svg%3E");
  pointer-events: none;
  z-index: 0;
}

/* ===== NAVIGATION ===== */
.nav {
  position: fixed; top:0; left:0; right:0;
  background: rgba(250,249,247,0.85);
  backdrop-filter: blur(20px);
  -webkit-backdrop-filter: blur(20px);
  border-bottom: 1px solid var(--border);
  z-index: 100;
  padding: 0 40px;
  height: 64px;
  display: flex;
  align-items: center;
  justify-content: space-between;
}
.nav-brand {
  display: flex; align-items: center; gap: 10px;
  text-decoration: none; color: var(--text);
}
.nav-logo {
  width: 32px; height: 32px;
  background: var(--accent);
  border-radius: 8px;
  display: flex; align-items: center; justify-content: center;
  color: #fff; font-weight: 700; font-size: 14px;
}
.nav-title {
  font-weight: 700; font-size: 18px;
  letter-spacing: -0.3px;
}
.nav-subtitle {
  font-size: 13px; color: var(--text-secondary);
  margin-left: 12px;
  padding-left: 12px;
  border-left: 1px solid var(--border);
}
.nav-version {
  font-family: 'JetBrains Mono', monospace;
  font-size: 11px;
  color: var(--text-tertiary);
  background: var(--bg-subtle);
  padding: 4px 10px;
  border-radius: 20px;
}

/* ===== MAIN LAYOUT ===== */
.main {
  position: relative; z-index: 1;
  max-width: 960px;
  margin: 0 auto;
  padding: 96px 24px 80px;
}

/* ===== HERO ===== */
.hero {
  text-align: center;
  padding: 40px 0 48px;
}
.hero-badge {
  display: inline-flex; align-items: center; gap: 6px;
  background: var(--accent-light);
  color: var(--accent);
  font-size: 12px; font-weight: 600;
  padding: 6px 14px;
  border-radius: 20px;
  margin-bottom: 20px;
  letter-spacing: 0.5px;
  text-transform: uppercase;
}
.hero-badge::before {
  content: ''; width: 6px; height: 6px;
  background: var(--accent);
  border-radius: 50%;
  animation: pulse-dot 2s ease-in-out infinite;
}
@keyframes pulse-dot {
  0%,100% { opacity: 1; transform: scale(1); }
  50% { opacity: 0.5; transform: scale(1.3); }
}
.hero h1 {
  font-size: 42px; font-weight: 700;
  letter-spacing: -1.5px;
  line-height: 1.15;
  margin-bottom: 14px;
  background: linear-gradient(135deg, var(--text) 0%, var(--text-secondary) 100%);
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
}
.hero p {
  font-size: 17px; color: var(--text-secondary);
  max-width: 560px; margin: 0 auto;
  line-height: 1.65;
}

/* ===== FEATURE CHIPS ===== */
.features {
  display: flex; justify-content: center; gap: 8px;
  flex-wrap: wrap;
  margin-top: 28px;
}
.feature-chip {
  display: inline-flex; align-items: center; gap: 6px;
  background: var(--bg-card);
  border: 1px solid var(--border);
  padding: 8px 16px;
  border-radius: 24px;
  font-size: 13px;
  color: var(--text-secondary);
  transition: all 0.2s;
}
.feature-chip:hover {
  border-color: var(--accent);
  color: var(--accent);
  transform: translateY(-1px);
  box-shadow: var(--shadow-sm);
}
.feature-chip svg { width:14px; height:14px; flex-shrink:0; }

/* ===== STEP INDICATOR ===== */
.steps-bar {
  display: flex;
  align-items: center;
  gap: 0;
  margin-bottom: 32px;
  padding: 0 20px;
}
.step-item {
  display: flex; align-items: center; gap: 10px;
  flex: 1;
  position: relative;
}
.step-item:not(:last-child)::after {
  content: '';
  flex: 1;
  height: 2px;
  background: var(--border);
  margin: 0 12px;
  transition: background 0.4s;
}
.step-item.active:not(:last-child)::after { background: var(--accent); }
.step-item.done:not(:last-child)::after { background: var(--green); }
.step-num {
  width: 32px; height: 32px;
  border-radius: 50%;
  border: 2px solid var(--border);
  display: flex; align-items: center; justify-content: center;
  font-size: 13px; font-weight: 600;
  color: var(--text-tertiary);
  background: var(--bg-card);
  flex-shrink: 0;
  transition: all 0.3s;
}
.step-item.active .step-num {
  border-color: var(--accent);
  color: var(--accent);
  background: var(--accent-light);
}
.step-item.done .step-num {
  border-color: var(--green);
  color: #fff;
  background: var(--green);
}
.step-label {
  font-size: 13px; font-weight: 500;
  color: var(--text-tertiary);
  white-space: nowrap;
}
.step-item.active .step-label { color: var(--accent); }
.step-item.done .step-label { color: var(--green); }

/* ===== CARDS ===== */
.card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: var(--radius-lg);
  padding: 32px;
  margin-bottom: 20px;
  box-shadow: var(--shadow-sm);
  transition: box-shadow 0.3s;
}
.card:hover { box-shadow: var(--shadow-md); }

.card-header {
  display: flex; align-items: center; gap: 12px;
  margin-bottom: 24px;
}
.card-icon {
  width: 40px; height: 40px;
  border-radius: 10px;
  display: flex; align-items: center; justify-content: center;
  font-size: 18px;
  flex-shrink: 0;
}
.card-icon.blue { background: var(--accent-light); color: var(--accent); }
.card-icon.amber { background: var(--yellow-bg); color: var(--yellow); }
.card-icon.teal { background: #F0FDFA; color: #0D9488; }
.card-title { font-size: 17px; font-weight: 600; letter-spacing: -0.3px; }
.card-desc { font-size: 13px; color: var(--text-secondary); margin-top: 2px; }

/* ===== FILE UPLOAD ===== */
.upload-grid {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 16px;
}
@media(max-width:640px) { .upload-grid { grid-template-columns: 1fr; } }

.upload-zone {
  border: 2px dashed var(--border);
  border-radius: var(--radius);
  padding: 32px 24px;
  text-align: center;
  cursor: pointer;
  transition: all 0.25s;
  position: relative;
  overflow: hidden;
}
.upload-zone:hover {
  border-color: var(--accent);
  background: #FAFBFF;
}
.upload-zone.dragover {
  border-color: var(--accent);
  background: var(--accent-light);
  transform: scale(1.01);
}
.upload-zone.has-file {
  border-color: var(--green);
  border-style: solid;
  background: var(--green-bg);
}
.upload-zone input[type="file"] {
  position: absolute;
  inset: 0;
  opacity: 0;
  cursor: pointer;
}
.upload-icon {
  width: 48px; height: 48px;
  margin: 0 auto 12px;
  border-radius: 12px;
  display: flex; align-items: center; justify-content: center;
  font-size: 22px;
}
.upload-zone:not(.has-file) .upload-icon {
  background: var(--bg-subtle);
  color: var(--text-tertiary);
}
.upload-zone.has-file .upload-icon {
  background: var(--green);
  color: #fff;
}
.upload-label {
  font-size: 14px; font-weight: 600;
  margin-bottom: 4px;
}
.upload-hint {
  font-size: 12px;
  color: var(--text-tertiary);
}
.upload-zone.has-file .upload-hint { color: var(--green); }
.file-tag {
  display: inline-flex; align-items: center; gap: 4px;
  background: var(--bg-subtle);
  padding: 3px 10px;
  border-radius: 12px;
  font-size: 11px;
  color: var(--text-secondary);
  margin-top: 8px;
  font-family: 'JetBrains Mono', monospace;
}
.upload-zone.has-file .file-tag {
  background: rgba(5,150,105,0.1);
  color: var(--green);
}

/* ===== SCAN INFO ===== */
.scan-info {
  margin-top: 16px;
  display: flex; gap: 12px;
}
.scan-badge {
  display: inline-flex; align-items: center; gap: 6px;
  padding: 6px 14px;
  border-radius: var(--radius-sm);
  font-size: 12px;
  font-weight: 500;
  background: var(--bg-subtle);
  color: var(--text-secondary);
  border: 1px solid var(--border);
}
.scan-badge.standard { background: var(--green-bg); color: var(--green); border-color: rgba(5,150,105,0.2); }
.scan-badge.multi { background: var(--yellow-bg); color: var(--yellow); border-color: rgba(217,119,6,0.2); }

/* ===== CONFIG ===== */
.config-grid {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 20px;
}
@media(max-width:640px) { .config-grid { grid-template-columns: 1fr; } }
.config-group label {
  display: block;
  font-size: 13px; font-weight: 600;
  margin-bottom: 8px;
  color: var(--text-secondary);
  text-transform: uppercase;
  letter-spacing: 0.5px;
}
.config-group select,
.config-group input[type="number"] {
  width: 100%;
  padding: 10px 14px;
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  font-size: 14px;
  font-family: inherit;
  background: var(--bg-card);
  color: var(--text);
  transition: border-color 0.2s;
  appearance: none;
  -webkit-appearance: none;
}
.config-group select {
  background-image: url("data:image/svg+xml,%3Csvg width='10' height='6' viewBox='0 0 10 6' xmlns='http://www.w3.org/2000/svg'%3E%3Cpath d='M1 1l4 4 4-4' stroke='%239C9792' stroke-width='1.5' fill='none' stroke-linecap='round' stroke-linejoin='round'/%3E%3C/svg%3E");
  background-repeat: no-repeat;
  background-position: right 14px center;
  padding-right: 36px;
}
.config-group select:focus,
.config-group input:focus {
  outline: none;
  border-color: var(--accent);
  box-shadow: 0 0 0 3px rgba(37,99,235,0.1);
}
.config-hint {
  font-size: 11px;
  color: var(--text-tertiary);
  margin-top: 6px;
}

/* ===== BUTTONS ===== */
.btn {
  display: inline-flex; align-items: center; justify-content: center; gap: 8px;
  padding: 12px 28px;
  border-radius: var(--radius-sm);
  font-size: 14px; font-weight: 600;
  font-family: inherit;
  border: none;
  cursor: pointer;
  transition: all 0.2s;
  text-decoration: none;
}
.btn-primary {
  background: var(--accent);
  color: #fff;
  box-shadow: 0 1px 3px rgba(37,99,235,0.3);
}
.btn-primary:hover:not(:disabled) {
  background: var(--accent-hover);
  transform: translateY(-1px);
  box-shadow: 0 4px 12px rgba(37,99,235,0.35);
}
.btn-primary:disabled {
  opacity: 0.45;
  cursor: not-allowed;
  transform: none;
}
.btn-secondary {
  background: var(--bg-subtle);
  color: var(--text);
  border: 1px solid var(--border);
}
.btn-secondary:hover { background: var(--bg-hover); }
.btn-lg { padding: 14px 36px; font-size: 15px; border-radius: var(--radius); }
.btn-block { width: 100%; }
.btn svg { width: 16px; height: 16px; }

/* ===== PROGRESS ===== */
.progress-container {
  text-align: center;
  padding: 20px 0;
}
.progress-ring-wrap {
  position: relative;
  width: 140px; height: 140px;
  margin: 0 auto 24px;
}
.progress-ring {
  width: 140px; height: 140px;
  transform: rotate(-90deg);
}
.progress-ring-bg {
  fill: none;
  stroke: var(--border);
  stroke-width: 6;
}
.progress-ring-fill {
  fill: none;
  stroke: var(--accent);
  stroke-width: 6;
  stroke-linecap: round;
  stroke-dasharray: 408;
  stroke-dashoffset: 408;
  transition: stroke-dashoffset 0.4s ease;
}
.progress-pct {
  position: absolute;
  inset: 0;
  display: flex; flex-direction: column;
  align-items: center; justify-content: center;
}
.progress-pct-num {
  font-size: 32px; font-weight: 700;
  letter-spacing: -1px;
  font-family: 'JetBrains Mono', monospace;
}
.progress-pct-label {
  font-size: 11px;
  color: var(--text-tertiary);
  text-transform: uppercase;
  letter-spacing: 1px;
}
.progress-status {
  font-size: 14px;
  color: var(--text-secondary);
  margin-bottom: 6px;
}
.progress-detail {
  font-size: 12px;
  color: var(--text-tertiary);
  font-family: 'JetBrains Mono', monospace;
}

/* ===== RESULTS ===== */
.stats-grid {
  display: grid;
  grid-template-columns: repeat(4, 1fr);
  gap: 12px;
  margin-bottom: 28px;
}
@media(max-width:640px) { .stats-grid { grid-template-columns: repeat(2, 1fr); } }
.stat-card {
  background: var(--bg-subtle);
  border-radius: var(--radius);
  padding: 18px;
  text-align: center;
}
.stat-value {
  font-size: 28px; font-weight: 700;
  letter-spacing: -1px;
  font-family: 'JetBrains Mono', monospace;
}
.stat-label {
  font-size: 12px;
  color: var(--text-secondary);
  margin-top: 4px;
}
.stat-card.green .stat-value { color: var(--green); }
.stat-card.yellow .stat-value { color: var(--yellow); }
.stat-card.red .stat-value { color: var(--red); }

/* ===== RESULT TABLE ===== */
.table-wrap {
  overflow-x: auto;
  border: 1px solid var(--border);
  border-radius: var(--radius);
  margin-top: 20px;
}
.result-table {
  width: 100%;
  border-collapse: collapse;
  font-size: 13px;
}
.result-table th {
  background: var(--bg-subtle);
  padding: 12px 16px;
  text-align: left;
  font-weight: 600;
  font-size: 12px;
  text-transform: uppercase;
  letter-spacing: 0.5px;
  color: var(--text-secondary);
  border-bottom: 1px solid var(--border);
  position: sticky; top: 0;
}
.result-table td {
  padding: 10px 16px;
  border-bottom: 1px solid var(--border);
  vertical-align: middle;
}
.result-table tr:last-child td { border-bottom: none; }
.result-table tr:hover td { background: var(--bg-subtle); }
.score-badge {
  display: inline-block;
  padding: 3px 10px;
  border-radius: 12px;
  font-size: 12px;
  font-weight: 600;
  font-family: 'JetBrains Mono', monospace;
}
.score-badge.high { background: var(--green-bg); color: var(--green); }
.score-badge.mid { background: var(--yellow-bg); color: var(--yellow); }
.score-badge.low { background: var(--red-bg); color: var(--red); }

.download-section {
  display: flex;
  align-items: center;
  gap: 16px;
  margin-top: 28px;
  padding: 20px 24px;
  background: var(--accent-light);
  border-radius: var(--radius);
  border: 1px solid rgba(37,99,235,0.15);
}
.download-info { flex: 1; }
.download-info h4 { font-size: 15px; font-weight: 600; margin-bottom: 2px; }
.download-info p { font-size: 12px; color: var(--text-secondary); }

/* ===== SECTION VISIBILITY ===== */
.section { display: none; }
.section.active { display: block; animation: fadeUp 0.4s ease; }
@keyframes fadeUp {
  from { opacity: 0; transform: translateY(12px); }
  to { opacity: 1; transform: translateY(0); }
}

/* ===== FOOTER ===== */
.footer {
  text-align: center;
  padding: 32px 0;
  font-size: 12px;
  color: var(--text-tertiary);
  border-top: 1px solid var(--border);
  margin-top: 48px;
}

/* ===== TOAST ===== */
.toast {
  position: fixed;
  bottom: 24px; right: 24px;
  background: var(--text);
  color: #fff;
  padding: 12px 20px;
  border-radius: var(--radius-sm);
  font-size: 13px;
  box-shadow: var(--shadow-lg);
  z-index: 1000;
  transform: translateY(80px);
  opacity: 0;
  transition: all 0.3s ease;
}
.toast.show { transform: translateY(0); opacity: 1; }

/* ===== RESET BTN ===== */
.reset-bar {
  display: flex;
  justify-content: center;
  margin-top: 24px;
}
</style>
</head>
<body>

<!-- NAV -->
<nav class="nav">
  <a class="nav-brand" href="/">
    <div class="nav-logo">NL</div>
    <span class="nav-title">NameLink</span>
    <span class="nav-subtitle">公司名称智能匹配系统</span>
  </a>
  <span class="nav-version">v3.0 Web</span>
</nav>

<main class="main">

  <!-- HERO -->
  <div class="hero">
    <div class="hero-badge">Smart Matching Engine</div>
    <h1>公司名称智能匹配</h1>
    <p>上传两张公司名单表格，系统将通过多维文本标准化与三重相似度算法，自动完成智能模糊匹配并输出专业报告</p>
    <div class="features">
      <div class="feature-chip">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 12l2 2 4-4"/><circle cx="12" cy="12" r="10"/></svg>
        繁简体 / 全半角标准化
      </div>
      <div class="feature-chip">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M4 6h16M4 12h16M4 18h7"/></svg>
        三重相似度打分
      </div>
      <div class="feature-chip">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 00-2 2v16a2 2 0 002 2h12a2 2 0 002-2V8z"/><path d="M14 2v6h6"/></svg>
        Excel 专业报告输出
      </div>
      <div class="feature-chip">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 2v4M12 18v4M4.93 4.93l2.83 2.83M16.24 16.24l2.83 2.83M2 12h4M18 12h4M4.93 19.07l2.83-2.83M16.24 7.76l2.83-2.83"/></svg>
        公司后缀智能统一
      </div>
    </div>
  </div>

  <!-- STEP INDICATOR -->
  <div class="steps-bar">
    <div class="step-item active" id="step1">
      <div class="step-num">1</div>
      <span class="step-label">上传文件</span>
    </div>
    <div class="step-item" id="step2">
      <div class="step-num">2</div>
      <span class="step-label">配置参数</span>
    </div>
    <div class="step-item" id="step3">
      <div class="step-num">3</div>
      <span class="step-label">匹配处理</span>
    </div>
    <div class="step-item" id="step4">
      <div class="step-num">4</div>
      <span class="step-label">查看结果</span>
    </div>
  </div>

  <!-- SECTION 1: UPLOAD -->
  <div class="section active" id="sec-upload">
    <div class="card">
      <div class="card-header">
        <div class="card-icon blue">
          <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
        </div>
        <div>
          <div class="card-title">上传公司名单</div>
          <div class="card-desc">支持 .xlsx, .xls, .csv 格式</div>
        </div>
      </div>

      <div class="upload-grid">
        <!-- A表 -->
        <div class="upload-zone" id="zone-a" onclick="document.getElementById('file-a').click()">
          <input type="file" id="file-a" accept=".xlsx,.xls,.csv" onchange="handleFileSelect('a', this)">
          <div class="upload-icon" id="icon-a">📋</div>
          <div class="upload-label" id="label-a">A表 — 正确公司名单</div>
          <div class="upload-hint" id="hint-a">点击或拖拽上传文件</div>
          <div class="file-tag" id="tag-a" style="display:none"></div>
        </div>
        <!-- B表 -->
        <div class="upload-zone" id="zone-b" onclick="document.getElementById('file-b').click()">
          <input type="file" id="file-b" accept=".xlsx,.xls,.csv" onchange="handleFileSelect('b', this)">
          <div class="upload-icon" id="icon-b">🔍</div>
          <div class="upload-label" id="label-b">B表 — 待核对公司名单</div>
          <div class="upload-hint" id="hint-b">点击或拖拽上传文件</div>
          <div class="file-tag" id="tag-b" style="display:none"></div>
        </div>
      </div>

      <!-- Scan info badges -->
      <div class="scan-info" id="scan-info" style="display:none">
        <div class="scan-badge" id="scan-a-badge"></div>
        <div class="scan-badge" id="scan-b-badge"></div>
      </div>

      <div style="margin-top:24px; text-align:center;">
        <button class="btn btn-primary btn-lg" id="btn-next-config" disabled onclick="goToConfig()">
          下一步：配置参数
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M5 12h14M12 5l7 7-7 7"/></svg>
        </button>
      </div>
    </div>
  </div>

  <!-- SECTION 2: CONFIG -->
  <div class="section" id="sec-config">
    <div class="card">
      <div class="card-header">
        <div class="card-icon amber">⚙️</div>
        <div>
          <div class="card-title">匹配参数设置</div>
          <div class="card-desc">根据需要调整参数，通常使用默认值即可获得最佳效果</div>
        </div>
      </div>

      <div class="config-grid">
        <div class="config-group">
          <label>A表读取模式</label>
          <select id="cfg-scan-a">
            <option value="auto">自动检测（推荐）</option>
            <option value="first">仅读取第一列</option>
            <option value="all">全表扫描所有单元格</option>
          </select>
          <div class="config-hint" id="cfg-scan-a-hint">系统将自动判断表格格式</div>
        </div>
        <div class="config-group">
          <label>B表读取模式</label>
          <select id="cfg-scan-b">
            <option value="auto">自动检测（推荐）</option>
            <option value="first">仅读取第一列</option>
            <option value="all">全表扫描所有单元格</option>
          </select>
          <div class="config-hint" id="cfg-scan-b-hint">系统将自动判断表格格式</div>
        </div>
        <div class="config-group">
          <label>候选明细百分比</label>
          <select id="cfg-pct">
            <option value="3">前 3%</option>
            <option value="5" selected>前 5%（推荐）</option>
            <option value="10">前 10%</option>
            <option value="15">前 15%</option>
            <option value="20">前 20%</option>
          </select>
          <div class="config-hint">在明细表中显示排名前 X% 的候选匹配公司</div>
        </div>
        <div class="config-group">
          <label>低匹配阈值</label>
          <select id="cfg-threshold">
            <option value="0.25">25%</option>
            <option value="0.35" selected>35%（推荐）</option>
            <option value="0.45">45%</option>
            <option value="0.55">55%</option>
          </select>
          <div class="config-hint">低于此阈值的匹配将被标记为"未找到"</div>
        </div>
      </div>

      <div style="margin-top:28px; display:flex; justify-content:center; gap:12px;">
        <button class="btn btn-secondary" onclick="goToUpload()">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="width:14px;height:14px"><path d="M19 12H5M12 19l-7-7 7-7"/></svg>
          返回
        </button>
        <button class="btn btn-primary btn-lg" onclick="startMatching()">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M5 3l14 9-14 9V3z"/></svg>
          开始匹配
        </button>
      </div>
    </div>
  </div>

  <!-- SECTION 3: PROGRESS -->
  <div class="section" id="sec-progress">
    <div class="card">
      <div class="card-header">
        <div class="card-icon teal">⚡</div>
        <div>
          <div class="card-title">正在匹配处理</div>
          <div class="card-desc" id="progress-subtitle">请稍候，正在分析数据...</div>
        </div>
      </div>
      <div class="progress-container">
        <div class="progress-ring-wrap">
          <svg class="progress-ring" viewBox="0 0 140 140">
            <circle class="progress-ring-bg" cx="70" cy="70" r="65"/>
            <circle class="progress-ring-fill" id="progress-circle" cx="70" cy="70" r="65"/>
          </svg>
          <div class="progress-pct">
            <div class="progress-pct-num" id="progress-num">0%</div>
            <div class="progress-pct-label">完成</div>
          </div>
        </div>
        <div class="progress-status" id="progress-msg">准备中...</div>
        <div class="progress-detail" id="progress-detail"></div>
      </div>
    </div>
  </div>

  <!-- SECTION 4: RESULTS -->
  <div class="section" id="sec-results">
    <div class="card">
      <div class="card-header">
        <div class="card-icon blue">📊</div>
        <div>
          <div class="card-title">匹配结果概览</div>
          <div class="card-desc" id="result-summary"></div>
        </div>
      </div>

      <div class="stats-grid" id="stats-grid"></div>

      <!-- Download -->
      <div class="download-section">
        <div class="download-info">
          <h4>📎 Excel 完整报告已生成</h4>
          <p>包含「最佳匹配结果」和「候选明细」两个工作表</p>
        </div>
        <a class="btn btn-primary" id="btn-download" href="#" download>
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
          下载报告
        </a>
      </div>

      <!-- Preview Table -->
      <div class="table-wrap" id="table-wrap" style="max-height:480px;overflow-y:auto;">
        <table class="result-table" id="result-table">
          <thead>
            <tr>
              <th style="width:5%">#</th>
              <th style="width:35%">B表公司</th>
              <th style="width:35%">A表最佳匹配</th>
              <th style="width:12%">匹配度</th>
              <th style="width:13%">状态</th>
            </tr>
          </thead>
          <tbody id="result-tbody"></tbody>
        </table>
      </div>

      <div class="reset-bar">
        <button class="btn btn-secondary" onclick="resetAll()">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="width:14px;height:14px"><path d="M1 4v6h6"/><path d="M3.51 15a9 9 0 105.64-11.36L1 10"/></svg>
          重新匹配
        </button>
      </div>
    </div>
  </div>

</main>

<footer class="footer">
  NameLink v3.0 Web — Company Name Intelligent Matching System<br>
  多维文本标准化 · 字符重叠率 · LCS序列相似度 · 分词重叠率
</footer>

<div class="toast" id="toast"></div>

<script>
// ===== STATE =====
let fileA = null, fileB = null;
let scanInfoA = null, scanInfoB = null;
let taskId = null;
let pollTimer = null;

// ===== DRAG & DROP =====
['zone-a','zone-b'].forEach(id => {
  const z = document.getElementById(id);
  z.addEventListener('dragover', e => { e.preventDefault(); z.classList.add('dragover'); });
  z.addEventListener('dragleave', () => z.classList.remove('dragover'));
  z.addEventListener('drop', e => {
    e.preventDefault(); z.classList.remove('dragover');
    const files = e.dataTransfer.files;
    if(files.length) {
      const which = id === 'zone-a' ? 'a' : 'b';
      const inp = document.getElementById('file-'+which);
      inp.files = files;
      handleFileSelect(which, inp);
    }
  });
});

function handleFileSelect(which, input) {
  const file = input.files[0];
  if(!file) return;
  const zone = document.getElementById('zone-' + which);
  const icon = document.getElementById('icon-' + which);
  const label = document.getElementById('label-' + which);
  const hint = document.getElementById('hint-' + which);
  const tag = document.getElementById('tag-' + which);

  zone.classList.add('has-file');
  icon.textContent = '✓';
  label.textContent = file.name;
  hint.textContent = '文件已选择';
  tag.style.display = 'inline-flex';
  tag.textContent = (file.size / 1024).toFixed(1) + ' KB';

  if(which === 'a') fileA = file;
  else fileB = file;

  checkReady();
  analyzeFile(which, file);
}

function checkReady() {
  document.getElementById('btn-next-config').disabled = !(fileA && fileB);
}

function analyzeFile(which, file) {
  const fd = new FormData();
  fd.append('file', file);
  fetch('/api/analyze', { method:'POST', body: fd })
    .then(r => r.json())
    .then(data => {
      if(which === 'a') scanInfoA = data;
      else scanInfoB = data;
      updateScanInfo();
    })
    .catch(() => {});
}

function updateScanInfo() {
  const cont = document.getElementById('scan-info');
  if(!scanInfoA && !scanInfoB) { cont.style.display = 'none'; return; }
  cont.style.display = 'flex';

  if(scanInfoA) {
    const b = document.getElementById('scan-a-badge');
    b.className = 'scan-badge ' + (scanInfoA.is_standard ? 'standard' : 'multi');
    b.textContent = 'A表: ' + (scanInfoA.is_standard ? '标准单列格式' : '多列格式') +
                    ' · ' + scanInfoA.count + ' 条';
  }
  if(scanInfoB) {
    const b = document.getElementById('scan-b-badge');
    b.className = 'scan-badge ' + (scanInfoB.is_standard ? 'standard' : 'multi');
    b.textContent = 'B表: ' + (scanInfoB.is_standard ? '标准单列格式' : '多列格式') +
                    ' · ' + scanInfoB.count + ' 条';
  }
}

// ===== NAVIGATION =====
function setStep(n) {
  for(let i=1;i<=4;i++) {
    const el = document.getElementById('step'+i);
    el.className = 'step-item' + (i < n ? ' done' : (i === n ? ' active' : ''));
  }
}
function showSection(id) {
  document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
  document.getElementById(id).classList.add('active');
}

function goToUpload() { setStep(1); showSection('sec-upload'); }
function goToConfig() { setStep(2); showSection('sec-config'); updateConfigHints(); }

function updateConfigHints() {
  if(scanInfoA) {
    const h = document.getElementById('cfg-scan-a-hint');
    h.textContent = scanInfoA.is_standard ?
      '检测结果：标准单列格式 (' + scanInfoA.count + ' 条)' :
      '检测结果：多列格式 (' + scanInfoA.count + ' 条文本)';
    if(!scanInfoA.is_standard) document.getElementById('cfg-scan-a').value = 'all';
  }
  if(scanInfoB) {
    const h = document.getElementById('cfg-scan-b-hint');
    h.textContent = scanInfoB.is_standard ?
      '检测结果：标准单列格式 (' + scanInfoB.count + ' 条)' :
      '检测结果：多列格式 (' + scanInfoB.count + ' 条文本)';
    if(!scanInfoB.is_standard) document.getElementById('cfg-scan-b').value = 'all';
  }
}

// ===== START MATCHING =====
function startMatching() {
  setStep(3); showSection('sec-progress');

  const fd = new FormData();
  fd.append('file_a', fileA);
  fd.append('file_b', fileB);

  const scanAVal = document.getElementById('cfg-scan-a').value;
  const scanBVal = document.getElementById('cfg-scan-b').value;
  let scanA, scanB;
  if(scanAVal === 'auto') scanA = scanInfoA ? !scanInfoA.is_standard : false;
  else scanA = scanAVal === 'all';
  if(scanBVal === 'auto') scanB = scanInfoB ? !scanInfoB.is_standard : false;
  else scanB = scanBVal === 'all';

  fd.append('scan_a', scanA ? '1' : '0');
  fd.append('scan_b', scanB ? '1' : '0');
  fd.append('top_pct', document.getElementById('cfg-pct').value);

  fetch('/api/start', { method:'POST', body: fd })
    .then(r => r.json())
    .then(data => {
      if(data.error) { showToast(data.error); return; }
      taskId = data.task_id;
      pollProgress();
    })
    .catch(e => showToast('上传失败: ' + e.message));
}

function pollProgress() {
  if(!taskId) return;
  fetch('/api/progress/' + taskId)
    .then(r => r.json())
    .then(data => {
      updateProgress(data);
      if(data.status === 'done') {
        setTimeout(() => showResults(data.result), 500);
      } else if(data.status === 'error') {
        showToast(data.message);
      } else {
        pollTimer = setTimeout(pollProgress, 300);
      }
    })
    .catch(() => { pollTimer = setTimeout(pollProgress, 1000); });
}

function updateProgress(data) {
  const pct = Math.round(data.progress || 0);
  document.getElementById('progress-num').textContent = pct + '%';
  document.getElementById('progress-msg').textContent = data.message || '';

  // Ring
  const circle = document.getElementById('progress-circle');
  const circumference = 2 * Math.PI * 65;
  circle.style.strokeDashoffset = circumference - (pct / 100) * circumference;

  let detail = '';
  if(data.status === 'matching' && data.current && data.total) {
    detail = data.current + ' / ' + data.total;
    if(data.current_name) detail += ' · ' + data.current_name;
  }
  if(data.count_a) detail = 'A表 ' + data.count_a + ' · B表 ' + (data.count_b||'?');
  document.getElementById('progress-detail').textContent = detail;

  const subtitle = document.getElementById('progress-subtitle');
  if(data.total_comparisons) {
    subtitle.textContent = '共计 ' + data.total_comparisons.toLocaleString() + ' 次比较';
  }
}

// ===== RESULTS =====
function showResults(result) {
  setStep(4); showSection('sec-results');

  document.getElementById('result-summary').textContent =
    '共匹配 ' + result.total + ' 家公司，其中高匹配 ' + result.high + ' 家';

  const grid = document.getElementById('stats-grid');
  grid.innerHTML = `
    <div class="stat-card"><div class="stat-value">${result.total}</div><div class="stat-label">总匹配数</div></div>
    <div class="stat-card green"><div class="stat-value">${result.high}</div><div class="stat-label">高匹配 ≥85%</div></div>
    <div class="stat-card yellow"><div class="stat-value">${result.mid}</div><div class="stat-label">中匹配 55-85%</div></div>
    <div class="stat-card red"><div class="stat-value">${result.low}</div><div class="stat-label">低匹配 &lt;55%</div></div>
  `;

  document.getElementById('btn-download').href = '/api/download/' + result.filename;

  const tbody = document.getElementById('result-tbody');
  tbody.innerHTML = '';
  result.preview.forEach((row, i) => {
    const statusMap = { high: '✓ 高匹配', mid: '⚠ 复核', low: '✗ 低匹配' };
    const tr = document.createElement('tr');
    tr.innerHTML = `
      <td style="color:var(--text-tertiary)">${i+1}</td>
      <td>${esc(row.b_name)}</td>
      <td>${esc(row.a_name)}</td>
      <td><span class="score-badge ${row.level}">${row.score}%</span></td>
      <td style="font-size:12px">${statusMap[row.level]}</td>
    `;
    tbody.appendChild(tr);
  });
  if(result.preview.length >= 100 && result.total > 100) {
    const tr = document.createElement('tr');
    tr.innerHTML = '<td colspan="5" style="text-align:center;color:var(--text-tertiary);font-size:12px;">... 显示前100条，完整结果请下载Excel报告 ...</td>';
    tbody.appendChild(tr);
  }
}

function esc(s) {
  const d = document.createElement('div');
  d.textContent = s;
  return d.innerHTML;
}

function resetAll() {
  fileA = null; fileB = null;
  scanInfoA = null; scanInfoB = null;
  taskId = null;
  if(pollTimer) clearTimeout(pollTimer);

  ['a','b'].forEach(w => {
    const zone = document.getElementById('zone-'+w);
    zone.classList.remove('has-file');
    document.getElementById('icon-'+w).textContent = w === 'a' ? '📋' : '🔍';
    document.getElementById('label-'+w).textContent = w === 'a' ? 'A表 — 正确公司名单' : 'B表 — 待核对公司名单';
    document.getElementById('hint-'+w).textContent = '点击或拖拽上传文件';
    document.getElementById('tag-'+w).style.display = 'none';
    document.getElementById('file-'+w).value = '';
  });
  document.getElementById('scan-info').style.display = 'none';
  document.getElementById('btn-next-config').disabled = true;

  // Reset progress ring
  document.getElementById('progress-circle').style.strokeDashoffset = 408;
  document.getElementById('progress-num').textContent = '0%';

  goToUpload();
}

function showToast(msg) {
  const t = document.getElementById('toast');
  t.textContent = msg;
  t.classList.add('show');
  setTimeout(() => t.classList.remove('show'), 3500);
}
</script>
</body>
</html>'''


# ======================== ROUTES ========================

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/api/analyze', methods=['POST'])
def analyze_file():
    """分析上传文件的格式"""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '未收到文件'}), 400

    filename = f"analyze_{uuid.uuid4().hex[:8]}_{file.filename}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        is_standard, total_rows, count = detect_table_format(filepath)
        return jsonify({
            'is_standard': is_standard,
            'total_rows': total_rows,
            'count': count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            os.remove(filepath)
        except:
            pass


@app.route('/api/start', methods=['POST'])
def start_matching():
    """启动匹配任务"""
    file_a = request.files.get('file_a')
    file_b = request.files.get('file_b')

    if not file_a or not file_b:
        return jsonify({'error': '请上传A表和B表'}), 400

    task_id = uuid.uuid4().hex[:12]

    # 保存文件
    path_a = os.path.join(UPLOAD_FOLDER, f"{task_id}_a_{file_a.filename}")
    path_b = os.path.join(UPLOAD_FOLDER, f"{task_id}_b_{file_b.filename}")
    file_a.save(path_a)
    file_b.save(path_b)

    scan_a = request.form.get('scan_a', '0') == '1'
    scan_b = request.form.get('scan_b', '0') == '1'
    top_pct = int(request.form.get('top_pct', '5')) / 100.0
    top_pct = max(0.01, min(1.0, top_pct))

    # 初始化任务状态
    tasks[task_id] = {
        'status': 'uploading',
        'progress': 0,
        'message': '文件上传完成，准备开始...',
        'current': 0,
        'total': 0,
        'count_a': 0,
        'count_b': 0,
        'current_name': '',
        'total_comparisons': 0,
        'result': None
    }

    # 启动后台线程
    thread = threading.Thread(
        target=run_matching_task,
        args=(task_id, path_a, path_b, scan_a, scan_b, top_pct),
        daemon=True
    )
    thread.start()

    return jsonify({'task_id': task_id})


@app.route('/api/progress/<task_id>')
def get_progress(task_id):
    """获取任务进度"""
    task = tasks.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    return jsonify(task)


@app.route('/api/download/<filename>')
def download_result(filename):
    """下载结果文件"""
    filepath = os.path.join(RESULT_FOLDER, filename)
    if not os.path.isfile(filepath):
        return "文件不存在", 404
    return send_file(filepath, as_attachment=True, download_name=filename)


# ======================== MAIN ========================
if __name__ == '__main__':
    print()
    print("=" * 60)
    print("  NameLink — 公司名称智能匹配系统 Web Edition")
    print("  Company Name Intelligent Matching System")
    print("=" * 60)
    print()
    print(f"  请在浏览器中打开: http://localhost:8080")
    print(f"  按 Ctrl+C 停止服务")
    print()
    app.run(host='0.0.0.0', port=8080, debug=False)