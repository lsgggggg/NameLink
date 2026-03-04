#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==========================================================================
  公司名称智能匹配系统 | Company Name Intelligent Matching System
==========================================================================
  用途: 将B表（待核对公司名单）与A表（正确公司名单）进行智能模糊匹配
  功能:
    1. 多维度文本标准化（大小写/繁简体/Ltd-Limited/空格/标点等）
    2. 字符级相似度 + 序列相似度 + 分词相似度 三重打分
    3. 输出最佳匹配表 + 前5%匹配明细表
    4. 交互式终端操作
==========================================================================
"""

import os
import sys
import re
import math
from collections import Counter

# ======================== 第三方库导入 ========================
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("[!] 正在安装 openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import pandas as pd
except ImportError:
    print("[!] 正在安装 pandas...")
    os.system(f"{sys.executable} -m pip install pandas -q")
    import pandas as pd

try:
    from opencc import OpenCC
    CC_T2S = OpenCC('t2s')  # 繁体转简体
    HAS_OPENCC = True
except ImportError:
    print("[i] opencc-python-reimplemented 未安装，尝试安装...")
    os.system(f"{sys.executable} -m pip install opencc-python-reimplemented -q")
    try:
        from opencc import OpenCC
        CC_T2S = OpenCC('t2s')
        HAS_OPENCC = True
    except ImportError:
        HAS_OPENCC = False
        print("[!] opencc 安装失败，繁简体转换功能将使用内置基础映射表")


# ======================== 常用繁简体映射（备用） ========================
TRAD_TO_SIMP = {
    '國': '国', '際': '际', '業': '业', '門': '门', '開': '开', '發': '发',
    '會': '会', '經': '经', '濟': '济', '貿': '贸', '電': '电', '機': '机',
    '設': '设', '計': '计', '環': '环', '達': '达', '運': '运', '輸': '输',
    '銀': '银', '號': '号', '萬': '万', '億': '亿', '華': '华', '東': '东',
    '車': '车', '馬': '马', '龍': '龙', '鳳': '凤', '聯': '联', '豐': '丰',
    '廣': '广', '產': '产', '實': '实', '寶': '宝', '與': '与', '對': '对',
    '長': '长', '創': '创', '風': '风', '園': '园', '築': '筑', '買': '买',
    '賣': '卖', '進': '进', '農': '农', '漁': '渔', '礦': '矿', '鋼': '钢',
    '鐵': '铁', '飛': '飞', '樂': '乐', '書': '书', '學': '学', '類': '类',
    '師': '师', '個': '个', '從': '从', '復': '复', '後': '后', '態': '态',
    '總': '总', '團': '团', '層': '层', '陽': '阳', '島': '岛', '區': '区',
    '織': '织', '紡': '纺', '線': '线', '練': '练', '組': '组', '終': '终',
    '約': '约', '級': '级', '結': '结', '給': '给', '統': '统', '網': '网',
    '護': '护', '報': '报', '場': '场', '塊': '块', '壓': '压', '夢': '梦',
    '備': '备', '處': '处', '複': '复', '裝': '装', '規': '规', '視': '视',
    '覺': '觉', '記': '记', '許': '许', '論': '论', '證': '证', '識': '识',
    '議': '议', '財': '财', '質': '质', '貨': '货', '資': '资', '賢': '贤',
    '軟': '软', '輝': '辉', '辦': '办', '邊': '边', '遠': '远', '選': '选',
    '鑫': '鑫', '關': '关', '陳': '陈', '項': '项', '順': '顺', '領': '领',
    '題': '题', '點': '点', '傳': '传', '僅': '仅', '優': '优', '衛': '卫',
    '寫': '写', '導': '导', '盡': '尽', '屬': '属', '嶺': '岭', '幣': '币',
    '廳': '厅', '張': '张', '彈': '弹', '強': '强', '歸': '归', '當': '当',
    '徑': '径', '從': '从', '應': '应', '戲': '戏', '擴': '扩', '據': '据',
    '損': '损', '搖': '摇', '構': '构', '標': '标', '歐': '欧', '歷': '历',
    '殘': '残', '滅': '灭', '瀾': '澜', '灣': '湾', '為': '为', '無': '无',
    '營': '营', '獨': '独', '獎': '奖', '現': '现', '瑣': '琐', '畫': '画',
    '異': '异', '療': '疗', '盤': '盘', '監': '监', '碼': '码', '積': '积',
    '稅': '税', '穩': '稳', '競': '竞', '節': '节', '範': '范', '築': '筑',
    '係': '系', '紀': '纪', '蘭': '兰', '蘇': '苏', '術': '术', '衝': '冲',
    '補': '补', '製': '制', '覆': '覆', '親': '亲', '觸': '触', '設': '设',
    '訊': '讯', '調': '调', '談': '谈', '請': '请', '識': '识', '變': '变',
    '讓': '让', '豬': '猪', '負': '负', '販': '贩', '費': '费', '離': '离',
    '雲': '云', '響': '响', '頭': '头', '顯': '显', '養': '养', '體': '体',
    '魚': '鱼', '麗': '丽', '齊': '齐', '齡': '龄',
}


# ======================== 文本标准化引擎 ========================

def traditional_to_simplified(text: str) -> str:
    """繁体转简体"""
    if HAS_OPENCC:
        return CC_T2S.convert(text)
    return ''.join(TRAD_TO_SIMP.get(c, c) for c in text)


# 公司后缀标准化映射
SUFFIX_NORMALIZE = [
    # 英文后缀统一
    (r'\blimited\b', 'ltd'),
    (r'\bltd\.?\b', 'ltd'),
    (r'\bcorporation\b', 'corp'),
    (r'\bcorp\.?\b', 'corp'),
    (r'\bcompany\b', 'co'),
    (r'\bco\.?\b', 'co'),
    (r'\bincorporated\b', 'inc'),
    (r'\binc\.?\b', 'inc'),
    (r'\bholdings?\b', 'holding'),
    (r'\bgroup\b', 'group'),
    (r'\binternational\b', 'intl'),
    (r'\bintl\.?\b', 'intl'),
    (r'\benterprise[s]?\b', 'enterprise'),
    (r'\binvestment[s]?\b', 'investment'),
    (r'\btechnolog(?:y|ies)\b', 'tech'),
    (r'\btech\.?\b', 'tech'),
    (r'\bmanagement\b', 'mgmt'),
    (r'\bmgmt\.?\b', 'mgmt'),
    (r'\bdevelopment\b', 'dev'),
    (r'\bdev\.?\b', 'dev'),
    (r'\bprivate\b', 'pvt'),
    (r'\bpvt\.?\b', 'pvt'),
    (r'\bsdn\.?\s*bhd\.?\b', 'sdn bhd'),
    (r'\bpte\.?\b', 'pte'),
    (r'\bb\.?\s*v\.?\b', 'bv'),
    (r'\bn\.?\s*v\.?\b', 'nv'),
    (r'\bg\.?\s*m\.?\s*b\.?\s*h\.?\b', 'gmbh'),
    (r'\bs\.?\s*a\.?\s*r\.?\s*l\.?\b', 'sarl'),
    (r'\bs\.?\s*\.?a\.?\b', 'sa'),
    # 中文后缀统一
    (r'有限责任公司', '有限公司'),
    (r'股份有限公司', '有限公司'),
    (r'（香港）', '(香港)'),
    (r'（中国）', '(中国)'),
    (r'（上海）', '(上海)'),
    (r'（北京）', '(北京)'),
    (r'（深圳）', '(深圳)'),
    (r'（广州）', '(广州)'),
]

# 要移除的无意义词
STOP_WORDS = {
    'the', 'of', 'and', '&', 'a', 'an', 'in', 'at', 'on', 'for', 'to', 'by',
    '-', '—', '–', '·', '•', ',', '.', '/', '\\',
}


def normalize_company_name(name: str) -> str:
    """
    多维度公司名称标准化：
    1. 去除首尾空白
    2. 繁体转简体
    3. 全角转半角
    4. 统一为小写
    5. 公司后缀标准化
    6. 统一所有括号为英文半角括号
    7. 统一标点符号（中文逗号句号等 -> 英文）
    8. 去除多余空格和标点
    """
    if not name or not isinstance(name, str):
        return ''

    text = name.strip()
    # 繁体 -> 简体
    text = traditional_to_simplified(text)
    # 全角 -> 半角
    text = full_to_half(text)
    # 小写
    text = text.lower()
    # 后缀标准化
    for pattern, replacement in SUFFIX_NORMALIZE:
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    # 统一所有括号为英文半角括号
    text = text.replace('（', '(').replace('）', ')')
    text = text.replace('【', '(').replace('】', ')')
    text = text.replace('〔', '(').replace('〕', ')')
    text = text.replace('［', '(').replace('］', ')')
    text = text.replace('｛', '(').replace('｝', ')')
    text = text.replace('{', '(').replace('}', ')')
    text = text.replace('[', '(').replace(']', ')')
    text = text.replace('﹙', '(').replace('﹚', ')')
    text = text.replace('《', '(').replace('》', ')')
    # 统一标点
    text = text.replace('，', ',').replace('。', '.').replace('；', ';')
    text = text.replace('：', ':').replace('、', ',').replace('～', '~')
    text = text.replace('\u3000', ' ')  # 中文全角空格
    text = text.replace('\xa0', ' ')    # non-breaking space
    text = text.replace('\t', ' ')      # tab
    text = text.replace('\r', ' ').replace('\n', ' ')  # 换行符
    # 去除多余空格（连续空格合并为一个）
    text = re.sub(r'\s+', ' ', text).strip()

    return text


def full_to_half(text: str) -> str:
    """全角字符转半角"""
    result = []
    for char in text:
        code = ord(char)
        if 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        elif code == 0x3000:
            result.append(' ')
        else:
            result.append(char)
    return ''.join(result)


def extract_chars(text: str) -> list:
    """将字符串拆分为字符列表（去除空格和标点后的纯内容字符）"""
    return [c for c in text if c.isalnum() or '\u4e00' <= c <= '\u9fff']


def extract_tokens(text: str) -> list:
    """
    分词：将字符串拆分为有意义的token列表
    中文：按单字切分
    英文：整个单词作为一个token + 单词内每个字母也作为独立token
          这样 "communication" vs "communications" 在字母级别也能匹配
    """
    tokens = []
    parts = text.split()
    for part in parts:
        if part.lower() in STOP_WORDS:
            continue
        i = 0
        current_eng = []
        while i < len(part):
            if '\u4e00' <= part[i] <= '\u9fff':
                if current_eng:
                    eng_word = ''.join(current_eng)
                    tokens.append(eng_word)
                    # 英文单词≥3个字母时，额外拆分每个字母
                    if len(eng_word) >= 3:
                        tokens.extend(list(eng_word))
                    current_eng = []
                tokens.append(part[i])
            elif part[i].isalnum():
                current_eng.append(part[i])
            else:
                if current_eng:
                    eng_word = ''.join(current_eng)
                    tokens.append(eng_word)
                    if len(eng_word) >= 3:
                        tokens.extend(list(eng_word))
                    current_eng = []
            i += 1
        if current_eng:
            eng_word = ''.join(current_eng)
            tokens.append(eng_word)
            if len(eng_word) >= 3:
                tokens.extend(list(eng_word))
    return tokens


# ======================== 相似度计算引擎 ========================

def char_overlap_ratio(chars_a: list, chars_b: list) -> float:
    """
    字符重叠率（基于字符频次的Jaccard风格）
    对B中每个字符，检查其在A中是否出现，计算命中比例
    双向取平均，避免长短串导致的偏差
    """
    if not chars_a or not chars_b:
        return 0.0

    counter_a = Counter(chars_a)
    counter_b = Counter(chars_b)

    # B中的字符在A中的命中数
    hit_b_in_a = sum(min(counter_b[c], counter_a.get(c, 0)) for c in counter_b)
    # A中的字符在B中的命中数
    hit_a_in_b = sum(min(counter_a[c], counter_b.get(c, 0)) for c in counter_a)

    ratio_b = hit_b_in_a / len(chars_b) if chars_b else 0
    ratio_a = hit_a_in_b / len(chars_a) if chars_a else 0

    # 加权平均，偏重B→A方向（因为B是待查的）
    return 0.6 * ratio_b + 0.4 * ratio_a


def longest_common_subsequence_ratio(s1: str, s2: str) -> float:
    """最长公共子序列比率 (LCS ratio)"""
    if not s1 or not s2:
        return 0.0
    m, n = len(s1), len(s2)
    # 优化：只需要两行
    prev = [0] * (n + 1)
    curr = [0] * (n + 1)
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if s1[i-1] == s2[j-1]:
                curr[j] = prev[j-1] + 1
            else:
                curr[j] = max(prev[j], curr[j-1])
        prev, curr = curr, [0] * (n + 1)
    lcs_len = prev[n]
    return (2.0 * lcs_len) / (m + n)


def token_overlap_ratio(tokens_a: list, tokens_b: list) -> float:
    """
    分词级别的重叠率（处理单词/字符级别的语义块匹配）
    """
    if not tokens_a or not tokens_b:
        return 0.0

    set_a = Counter(tokens_a)
    set_b = Counter(tokens_b)

    hit = sum(min(set_a.get(t, 0), set_b[t]) for t in set_b)
    total = max(len(tokens_a), len(tokens_b))
    return hit / total if total > 0 else 0.0


def compute_similarity(name_a_norm: str, name_b_norm: str,
                       chars_a: list, chars_b: list,
                       tokens_a: list, tokens_b: list) -> float:
    """
    综合相似度 = 加权组合三种相似度
      - 字符重叠率 (40%)：快速捕捉字符级别的相似
      - LCS比率 (35%)：捕捉序列顺序相似
      - 分词重叠率 (25%)：捕捉语义块匹配
    """
    if name_a_norm == name_b_norm:
        return 1.0

    s1 = char_overlap_ratio(chars_a, chars_b)
    s2 = longest_common_subsequence_ratio(name_a_norm, name_b_norm)
    s3 = token_overlap_ratio(tokens_a, tokens_b)

    return 0.40 * s1 + 0.35 * s2 + 0.25 * s3


# ======================== 核心匹配流程 ========================

def read_company_list(filepath: str, scan_all_cells: bool = False) -> list:
    """
    从Excel或CSV读取公司名称列表
    scan_all_cells=False: 标准模式，只读第一列（A列）
    scan_all_cells=True:  全表扫描模式，遍历所有单元格提取非空文本
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        df = pd.read_excel(filepath, header=None, dtype=str)
    elif ext == '.csv':
        df = pd.read_csv(filepath, header=None, dtype=str)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，请使用 .xlsx, .xls 或 .csv")

    if scan_all_cells:
        # 全表扫描：遍历所有列所有行，提取非空字符串，去重
        names_set = []
        seen = set()
        for col in df.columns:
            for val in df[col].dropna().astype(str):
                val = val.strip()
                if val and val not in seen:
                    # 仅过滤纯数字（含小数点和负号），任何含文字的内容全部保留，不漏掉任何一个
                    if not re.fullmatch(r'[-+]?\d*\.?\d+', val):
                        names_set.append(val)
                        seen.add(val)
        return names_set
    else:
        # 标准模式：取第一列
        names = df.iloc[:, 0].dropna().astype(str).tolist()
        names = [n.strip() for n in names if n.strip()]
        return names


def detect_table_format(filepath: str) -> bool:
    """
    自动检测表格是否为标准单列格式
    如果只有1列有数据，或第一列非空率>80%且其他列非空率<20%，视为标准格式
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xls'):
        df = pd.read_excel(filepath, header=None, dtype=str)
    elif ext == '.csv':
        df = pd.read_csv(filepath, header=None, dtype=str)
    else:
        return True

    if df.shape[1] <= 1:
        return True

    # 计算每列的非空率
    total_rows = len(df)
    if total_rows == 0:
        return True

    col_fill_rates = []
    for col in df.columns:
        non_empty = df[col].dropna().astype(str).apply(lambda x: len(x.strip()) > 0).sum()
        col_fill_rates.append(non_empty / total_rows)

    first_col_rate = col_fill_rates[0]
    other_cols_rate = sum(col_fill_rates[1:]) / max(len(col_fill_rates) - 1, 1)

    # 第一列填充率高且其他列低 → 标准单列格式
    if first_col_rate > 0.8 and other_cols_rate < 0.2:
        return True
    # 多列都有数据 → 混乱格式
    if sum(1 for r in col_fill_rates if r > 0.3) > 1:
        return False
    return True


def build_index(names: list) -> list:
    """
    为公司名称列表构建预处理索引
    返回: [(原始名称, 标准化名称, 字符列表, token列表), ...]
    """
    index = []
    for name in names:
        norm = normalize_company_name(name)
        chars = extract_chars(norm)
        tokens = extract_tokens(norm)
        index.append((name, norm, chars, tokens))
    return index


def match_all(index_a: list, index_b: list, top_pct: float = 0.05):
    """
    对B中每个公司，与A中所有公司计算相似度
    返回:
      best_matches: [(b_name, best_a_name, best_score), ...]
      detail_matches: [(b_name, [(a_name, score), ...]), ...]  (前top_pct%)
    """
    total_a = len(index_a)
    # 前5%对应的最低排名位置（按并列名次计算）
    top_rank_limit = max(1, math.ceil(total_a * top_pct))

    best_matches = []
    detail_matches = []

    total_b = len(index_b)

    for idx_b, (b_name, b_norm, b_chars, b_tokens) in enumerate(index_b):
        # 进度提示
        if (idx_b + 1) % 10 == 0 or idx_b == 0 or idx_b == total_b - 1:
            print(f"\r  匹配进度: {idx_b + 1}/{total_b} ({(idx_b+1)/total_b*100:.1f}%)", end='', flush=True)

        scores = []
        for (a_name, a_norm, a_chars, a_tokens) in index_a:
            sim = compute_similarity(a_norm, b_norm, a_chars, b_chars, a_tokens, b_tokens)
            scores.append((a_name, sim))

        # 按相似度降序排序
        scores.sort(key=lambda x: -x[1])

        # 最佳匹配
        best_a_name, best_score = scores[0] if scores else ('', 0.0)
        best_matches.append((b_name, best_a_name, best_score))

        # 计算并列名次下的前5%
        # 规则：相同分数的公司共享同一名次
        top_list = []
        current_rank = 0
        prev_score = None
        count = 0
        for a_name, score in scores:
            count += 1
            if score != prev_score:
                current_rank = count  # 并列名次：新分数时名次=当前序号
                prev_score = score
            if current_rank <= top_rank_limit:
                top_list.append((a_name, score))
            else:
                break

        detail_matches.append((b_name, top_list))

    print()  # 换行
    return best_matches, detail_matches


# ======================== Excel 输出 ========================

def write_best_match_sheet(wb: openpyxl.Workbook, best_matches: list, detail_matches: list, threshold: float = 0.35):
    """
    写入最佳匹配结果表
    列: B表公司 | A表最佳匹配 | 匹配度 | 备注 | 前X%候选公司(非100%时)
    """
    ws = wb.active
    ws.title = "最佳匹配结果"

    # 构建B公司名 -> 前X%候选列表的查找字典
    detail_lookup = {}
    for b_name, top_list in detail_matches:
        detail_lookup[b_name] = top_list

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["B表原始公司名称", "A表最佳匹配公司", "匹配度", "风险提示", "前X%候选公司明细"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 条件格式颜色
    green_fill = PatternFill("solid", fgColor="C6EFCE")   # 高匹配
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")   # 中等匹配
    red_fill = PatternFill("solid", fgColor="FFC7CE")      # 低匹配

    for row_idx, (b_name, a_name, score) in enumerate(best_matches, 2):
        ws.cell(row=row_idx, column=1, value=b_name).border = thin_border
        ws.cell(row=row_idx, column=2, value=a_name).border = thin_border

        score_cell = ws.cell(row=row_idx, column=3, value=round(score * 100, 2))
        score_cell.number_format = '0.00"%"'
        score_cell.border = thin_border
        score_cell.alignment = Alignment(horizontal="center")

        note_cell = ws.cell(row=row_idx, column=4)
        note_cell.border = thin_border

        if score >= 0.85:
            score_cell.fill = green_fill
            note_cell.value = "✓ 高度匹配"
            note_cell.font = Font(color="006100")
        elif score >= 0.55:
            score_cell.fill = yellow_fill
            note_cell.value = "⚠ 中等匹配，建议人工复核"
            note_cell.font = Font(color="9C6500")
        elif score >= threshold:
            score_cell.fill = red_fill
            note_cell.value = "⚠ 低匹配度，很可能不是同一公司，请重点核查"
            note_cell.font = Font(color="9C0006")
        else:
            score_cell.fill = red_fill
            note_cell.value = "✗ 在A表中未找到相似公司，可能A表中不存在该公司"
            note_cell.font = Font(color="9C0006", bold=True)

        # 第5列：非100%匹配时，列出前X%所有候选公司
        candidate_cell = ws.cell(row=row_idx, column=5)
        candidate_cell.border = thin_border
        candidate_cell.alignment = Alignment(vertical="top", wrap_text=True)

        if score < 1.0:
            top_list = detail_lookup.get(b_name, [])
            if top_list:
                lines = []
                rank = 0
                prev_s = None
                for a_candidate, s in top_list:
                    if s != prev_s:
                        rank += 1
                        prev_s = s
                    lines.append(f"#{rank} [{s*100:.1f}%] {a_candidate}")
                candidate_cell.value = "\n".join(lines)
                candidate_cell.font = Font(size=9, name="Arial")

    # 列宽
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 65

    # 冻结首行
    ws.freeze_panes = 'A2'


def write_detail_sheet(wb: openpyxl.Workbook, detail_matches: list, top_pct: float):
    """
    写入前5%匹配明细表
    每个B公司一个区块，列出匹配度前5%的所有A公司
    """
    ws = wb.create_sheet(title=f"前{int(top_pct*100)}%匹配明细")

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    section_font = Font(bold=True, size=11, color="2F5496", name="Arial")
    section_fill = PatternFill("solid", fgColor="D6E4F0")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 总表头
    main_headers = ["B表公司", "排名", "A表匹配公司", "匹配度(%)"]
    for col, h in enumerate(main_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    current_row = 2
    for b_name, top_list in detail_matches:
        # B公司名称标记
        b_cell = ws.cell(row=current_row, column=1, value=b_name)
        b_cell.font = section_font
        b_cell.fill = section_fill
        b_cell.border = thin_border

        if not top_list:
            ws.cell(row=current_row, column=2, value="-").border = thin_border
            ws.cell(row=current_row, column=3, value="无匹配结果").border = thin_border
            ws.cell(row=current_row, column=4, value=0).border = thin_border
            current_row += 1
            continue

        # 写入每个匹配结果
        rank = 0
        prev_score = None
        for i, (a_name, score) in enumerate(top_list):
            if score != prev_score:
                rank = i + 1
                prev_score = score

            row = current_row + i
            if i == 0:
                pass  # B公司名已写
            else:
                ws.cell(row=row, column=1, value="").border = thin_border

            ws.cell(row=row, column=2, value=rank).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=a_name).border = thin_border
            score_cell = ws.cell(row=row, column=4, value=round(score * 100, 2))
            score_cell.number_format = '0.00'
            score_cell.border = thin_border
            score_cell.alignment = Alignment(horizontal="center")

        current_row += len(top_list)

    # 列宽
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 14

    # 冻结首行
    ws.freeze_panes = 'A2'


def save_results(best_matches, detail_matches, output_path, top_pct):
    """保存结果到Excel"""
    wb = openpyxl.Workbook()
    write_best_match_sheet(wb, best_matches, detail_matches)
    write_detail_sheet(wb, detail_matches, top_pct)
    wb.save(output_path)
    print(f"\n  ✓ 结果已保存至: {output_path}")


# ======================== 交互式主程序 ========================

def main():
    print("=" * 68)
    print("    公司名称智能匹配系统 v3.0")
    print("    Company Name Intelligent Matching System")
    print("=" * 68)
    print()

    # 输入A表路径
    while True:
        path_a = input("  请输入A表（正确公司名单）文件路径: ").strip().strip('"').strip("'")
        if os.path.isfile(path_a):
            break
        print(f"  [!] 文件不存在: {path_a}，请重新输入\n")

    # 输入B表路径
    while True:
        path_b = input("  请输入B表（待核对公司名单）文件路径: ").strip().strip('"').strip("'")
        if os.path.isfile(path_b):
            break
        print(f"  [!] 文件不存在: {path_b}，请重新输入\n")

    # 输出路径
    default_output = os.path.join(os.path.dirname(path_b) or '.', "匹配结果.xlsx")
    output_path = input(f"  请输入输出文件路径 (回车默认: {default_output}): ").strip().strip('"').strip("'")
    if not output_path:
        output_path = default_output

    # 前百分比
    pct_input = input("  请输入要查看的前百分比 (回车默认5%): ").strip().replace('%', '')
    try:
        top_pct = float(pct_input) / 100.0 if pct_input else 0.05
    except ValueError:
        top_pct = 0.05
    top_pct = max(0.01, min(1.0, top_pct))

    # 自动检测表格格式
    print()
    scan_a = False
    scan_b = False

    is_a_standard = detect_table_format(path_a)
    if not is_a_standard:
        print("  [!] 检测到A表可能不是标准单列格式（数据分布在多列中）")
        choice = input("      是否扫描A表所有单元格？(y=全表扫描 / 回车=仅读第一列): ").strip().lower()
        scan_a = choice in ('y', 'yes', '是')
    else:
        print("  [✓] A表检测为标准单列格式，将读取第一列")

    is_b_standard = detect_table_format(path_b)
    if not is_b_standard:
        print("  [!] 检测到B表可能不是标准单列格式（数据分布在多列中）")
        choice = input("      是否扫描B表所有单元格？(y=全表扫描 / 回车=仅读第一列): ").strip().lower()
        scan_b = choice in ('y', 'yes', '是')
    else:
        print("  [✓] B表检测为标准单列格式，将读取第一列")

    print()
    print("  " + "-" * 50)
    print(f"  A表路径: {path_a}" + (" [全表扫描]" if scan_a else ""))
    print(f"  B表路径: {path_b}" + (" [全表扫描]" if scan_b else ""))
    print(f"  输出路径: {output_path}")
    print(f"  匹配明细: 前 {top_pct*100:.0f}%")
    print("  " + "-" * 50)
    print()

    # 读取数据
    print("  [1/4] 读取A表...")
    list_a = read_company_list(path_a, scan_all_cells=scan_a)
    print(f"        读取到 {len(list_a)} 家公司")

    print("  [2/4] 读取B表...")
    list_b = read_company_list(path_b, scan_all_cells=scan_b)
    print(f"        读取到 {len(list_b)} 家公司")

    # 构建索引
    print("  [3/4] 构建预处理索引...")
    index_a = build_index(list_a)
    index_b = build_index(list_b)
    print(f"        索引构建完成")

    # 匹配
    print(f"  [4/4] 执行匹配 ({len(list_b)} × {len(list_a)} = {len(list_b)*len(list_a):,} 次比较)...")
    best_matches, detail_matches = match_all(index_a, index_b, top_pct)

    # 保存
    save_results(best_matches, detail_matches, output_path, top_pct)

    # 简要统计
    high = sum(1 for _, _, s in best_matches if s >= 0.85)
    mid = sum(1 for _, _, s in best_matches if 0.55 <= s < 0.85)
    low = sum(1 for _, _, s in best_matches if s < 0.55)
    print()
    print("  ========== 匹配统计 ==========")
    print(f"  高匹配 (≥85%):  {high} 家")
    print(f"  中匹配 (55-85%): {mid} 家  ← 建议人工复核")
    print(f"  低匹配 (<55%):   {low} 家  ← 可能A表中不存在")
    print(f"  总计:            {len(best_matches)} 家")
    print("  " + "=" * 32)
    print()
    print("  完成！请打开输出的Excel文件查看详细匹配结果。")
    print()


if __name__ == '__main__':
    main()